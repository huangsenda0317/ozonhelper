"""新品刊登服务"""

from __future__ import annotations

import io
import json
import uuid
from decimal import Decimal

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from src.models.tracking_sync import ListingItem, ListingJob
from src.models.store import Store

REQUIRED_COLUMNS = ['offer_id', 'name', 'category_id', 'price', 'primary_image_url']
OPTIONAL_COLUMNS = ['attributes_json']


def build_listing_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'listing'
    ws.append(REQUIRED_COLUMNS + OPTIONAL_COLUMNS)
    ws.append(['SKU001', '示例商品', '17036168', '999.00', 'https://example.com/img.jpg', '{}'])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_listing_excel(content: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError('Excel 文件为空')
    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        raise ValueError(f'缺少必填列: {", ".join(missing)}')
    items: list[dict] = []
    for idx, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue
        data = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        offer_id = str(data.get('offer_id') or '').strip()
        if not offer_id:
            raise ValueError(f'第 {idx} 行 offer_id 为空')
        attrs = data.get('attributes_json')
        if isinstance(attrs, str) and attrs.strip():
            try:
                attrs = json.loads(attrs)
            except json.JSONDecodeError as exc:
                raise ValueError(f'第 {idx} 行 attributes_json 无效') from exc
        items.append(
            {
                'offer_id': offer_id,
                'name': str(data.get('name') or ''),
                'category_id': str(data.get('category_id') or ''),
                'price': Decimal(str(data.get('price') or 0)),
                'primary_image_url': str(data.get('primary_image_url') or ''),
                'attributes_json': attrs if isinstance(attrs, dict) else None,
            }
        )
    if len(items) > 500:
        raise ValueError('单次上传最多 500 行')
    return items


async def create_listing_job(
    db: AsyncSession,
    store: Store,
    filename: str,
    parsed_items: list[dict],
) -> ListingJob:
    job = ListingJob(
        store_id=store.id,
        filename=filename,
        status='pending',
        total_items=len(parsed_items),
    )
    db.add(job)
    await db.flush()
    for item in parsed_items:
        db.add(
            ListingItem(
                job_id=job.id,
                offer_id=item['offer_id'],
                name=item['name'],
                category_id=item.get('category_id'),
                price=item.get('price'),
                primary_image_url=item.get('primary_image_url'),
                attributes_json=item.get('attributes_json'),
            )
        )
    await db.flush()
    return job


async def list_listing_jobs(db: AsyncSession, store_id: uuid.UUID) -> list[ListingJob]:
    stmt = select(ListingJob).where(ListingJob.store_id == store_id).order_by(ListingJob.created_at.desc())
    return list((await db.execute(stmt)).scalars().all())


async def get_listing_job(db: AsyncSession, store_id: uuid.UUID, job_id: uuid.UUID) -> ListingJob | None:
    stmt = select(ListingJob).where(ListingJob.id == job_id, ListingJob.store_id == store_id)
    return (await db.execute(stmt)).scalar_one_or_none()


async def get_listing_items(db: AsyncSession, job_id: uuid.UUID) -> list[ListingItem]:
    stmt = select(ListingItem).where(ListingItem.job_id == job_id).order_by(ListingItem.created_at.asc())
    return list((await db.execute(stmt)).scalars().all())
